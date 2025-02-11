import pandas as pd
import sys
import io
import json
import chardet
from datetime import timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import Tk, filedialog

class FileProcessor:
    def __init__(self):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
        self.input_file = sys.argv[1]
        self.standard_times_file = sys.argv[2]
        self.score_df = None
        self.standard_times_df = None
        self.merged_df = None

    def detect_encoding(self, file_path):
        with open(file_path, 'rb') as file:
            result = chardet.detect(file.read(10000))
            return result['encoding']

    def add_standard_time(self, start_time_str, standard_time):
        start_time = pd.to_datetime(start_time_str, errors='coerce')
        if pd.isna(start_time) or pd.isna(standard_time):
            return ''
        seconds = self.time_to_seconds(standard_time)
        end_time = start_time + timedelta(seconds=seconds)
        return end_time.strftime('%Y/%m/%d %H:%M:%S')

    def read_files(self):
        try:
            input_file_encoding = self.detect_encoding(self.input_file)
            self.score_df = pd.read_csv(self.input_file, encoding=input_file_encoding)
            self.standard_times_df = pd.read_csv(self.standard_times_file, encoding='utf-8')
        except Exception as e:
            self.return_error_json(f"Error reading files: {e}")

    def process_data(self):
        self.score_df.columns = self.score_df.columns.str.strip()
        self.standard_times_df.columns = self.standard_times_df.columns.str.strip()
        self.score_df = self.score_df.apply(lambda x: x.map(lambda y: y.strip() if isinstance(y, str) else y))
        self.standard_times_df = self.standard_times_df.apply(lambda x: x.map(lambda y: y.strip() if isinstance(y, str) else y))
        
        self.score_df['所要時間（秒）'] = self.score_df['所要時間'].apply(self.time_to_seconds)
        
        self.merged_df = pd.merge(self.score_df, self.standard_times_df, on='コンテンツ名', how='left')
        self.merged_df['フォルダ名'] = self.score_df['フォルダ名']  # フォルダ名列を保持

        # 標準学習時間CSVに存在しないコンテンツ名を削除
        valid_content_names = self.standard_times_df['コンテンツ名'].unique()
        self.merged_df = self.merged_df[self.merged_df['コンテンツ名'].isin(valid_content_names)]

        self.merged_df['学習開始日時'] = pd.to_datetime(self.merged_df['学習開始日時'], errors='coerce')
        
        self.calculate_confirm_time_and_mark()

        # 学習完了日を計算
        self.merged_df['学習完了日'] = self.merged_df.apply(lambda row: self.add_standard_time(row['学習開始日時'], row['標準学習時間']) if row['マーク'] == 'O' else '', axis=1)

        self.merged_df = self.merged_df.groupby(['氏名', 'コンテンツ名'], as_index=False).agg({
            'グループ': 'first',
            'グループ(全階層)': 'first',
            'フォルダ名': 'first',
            '学習開始日時': 'first',
            '学習完了日': 'first',
            '所要時間': 'first',
            '標準学習時間': 'first',
            '確認時間': 'first',
            'マーク': 'first',
            '足りない時間': 'first',
            'URL': 'first',
            '所要時間（秒）': 'sum',  # 所要時間は合計する
        })
    
    def calculate_confirm_time_and_mark(self):
        grouped = self.merged_df.groupby(['氏名', 'コンテンツ名']).agg({
            '所要時間（秒）': 'sum',
            '標準学習時間': 'first'
        }).reset_index()

        for idx, row in grouped.iterrows():
            total_time = row['所要時間（秒）']
            standard_time = self.time_to_seconds(row['標準学習時間'])
            mark = 'O' if total_time >= standard_time else 'X'
            
            sub_df = self.merged_df[(self.merged_df['氏名'] == row['氏名']) & 
                                    (self.merged_df['コンテンツ名'] == row['コンテンツ名'])]
            
            if '理解度テスト' in row['コンテンツ名']:
                confirm_time = min(total_time, standard_time)
                self.merged_df.loc[sub_df.index, '確認時間'] = self.seconds_to_time(confirm_time)
                self.merged_df.loc[sub_df.index, 'マーク'] = 'O'
            else:
                if mark == 'O':
                    # 標準学習時間をhh:mm:ss形式に変換
                    h, m, s = map(int, row['標準学習時間'].split(':'))
                    self.merged_df.loc[sub_df.index, '確認時間'] = f"{h:02}:{m:02}:{s:02}"
                else:
                    self.merged_df.loc[sub_df.index, '確認時間'] = '00:00:00'
                self.merged_df.loc[sub_df.index, 'マーク'] = mark

            if mark == 'X':
                total_time_sum = sub_df['所要時間（秒）'].sum()
                missing_time = max(0, standard_time - total_time_sum)  # 足りない時間を計算

                # 足りない時間が発生する場合の処理
                if missing_time > 0:
                    h = int(missing_time // 3600)
                    m = int((missing_time % 3600) // 60)
                    s = int(missing_time % 60)
                    self.merged_df.loc[sub_df.index, '足りない時間'] = f"{h:02}:{m:02}:{s:02}"
                else:
                    self.merged_df.loc[sub_df.index, '足りない時間'] = '00:00:00'

                # 2行目以降の足りない時間は空欄に
                self.merged_df.loc[sub_df.index[1:], '足りない時間'] = ''

    def save_to_excel(self):
        try:
            root = Tk()
            root.withdraw()
            save_path = filedialog.asksaveasfilename(
                title="保存先を選択してください",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")]
            )

            if save_path:
                columns_to_output = [
                    'グループ', 'グループ(全階層)', '氏名', 'フォルダ名', 'コンテンツ名',
                    '学習開始日時', '学習完了日', '所要時間', '標準学習時間', '確認時間', 
                    'マーク', '足りない時間', 'URL'
                ]

                # 存在しない列を確認して追加
                for col in columns_to_output:
                    if col not in self.merged_df.columns:
                        self.merged_df[col] = ''  # 存在しない列は空文字で初期化

                # 理解度テストの足りない時間を非表示にする
                self.merged_df.loc[self.merged_df['コンテンツ名'].str.contains('理解度テスト', na=False), '足りない時間'] = ''

                # 時間形式のフォーマットを修正
                def format_time(time_str):
                    if pd.isna(time_str) or time_str == '':
                        return '00:00:00'
                    h, m, s = map(int, time_str.split(':'))
                    return f"{h:02}:{m:02}:{s:02}"

                # 所要時間、標準学習時間、確認時間のフォーマットを修正
                self.merged_df['所要時間'] = self.merged_df['所要時間'].apply(format_time)
                self.merged_df['標準学習時間'] = self.merged_df['標準学習時間'].apply(format_time)
                self.merged_df['確認時間'] = self.merged_df['確認時間'].apply(format_time)
                # 足りない時間が空でない場合のみフォーマットを修正
                self.merged_df.loc[self.merged_df['足りない時間'] != '', '足りない時間'] = \
                    self.merged_df.loc[self.merged_df['足りない時間'] != '', '足りない時間'].apply(format_time)
                
                # 足りない時間が00:00:00の場合は非表示にする
                self.merged_df.loc[self.merged_df['足りない時間'] == '00:00:00', '足りない時間'] = ''

                # 指定のカラムのみを出力
                self.merged_df = self.merged_df[columns_to_output]

                # コンテンツ名の順序を standard_times_df の順序に揃える
                content_order = self.standard_times_df['コンテンツ名'].drop_duplicates().dropna().tolist()
                self.merged_df['コンテンツ順序'] = pd.Categorical(self.merged_df['コンテンツ名'], categories=content_order, ordered=True)
                self.merged_df = self.merged_df.sort_values(['氏名', 'コンテンツ順序', '学習開始日時']).drop(columns=['コンテンツ順序'])

                # 修了証の数と合計時間をカウントして、ユーザーごとにソート
                user_stats = {}
                for name, group in self.merged_df.groupby('氏名'):
                    # フォルダ名ごとにマークXの有無を確認
                    folder_has_x_mark = {}
                    folder_has_completion = {}
                    
                    # フォルダ名ごとにマークXの有無と修了証の有無を確認
                    for _, row in group.iterrows():
                        folder_name = row['フォルダ名']
                        mark = row['マーク']
                        content_name = row['コンテンツ名']
                        
                        if folder_name:
                            if not folder_name in folder_has_x_mark:
                                folder_has_x_mark[folder_name] = False
                                folder_has_completion[folder_name] = False
                            if mark == 'X':
                                folder_has_x_mark[folder_name] = True
                            if '修了証' in str(content_name):
                                folder_has_completion[folder_name] = True

                    # オレンジ背景を除いた合計時間を計算
                    total_seconds = 0
                    for _, row in group.iterrows():
                        folder_name = row['フォルダ名']
                        should_be_orange = folder_name and (folder_has_x_mark[folder_name] or not folder_has_completion[folder_name])
                        if not should_be_orange and row['確認時間']:  # オレンジ背景でない場合のみ
                            total_seconds += self.time_to_seconds(row['確認時間'])

                    completion_count = group['コンテンツ名'].str.count('修了証').sum()
                    user_stats[name] = (completion_count, total_seconds)

                # 修了証の数が多い順、同数の場合は合計時間が多い順にソート
                sorted_users = sorted(user_stats.items(), key=lambda x: (-x[1][0], -x[1][1]))

                # 修了証の多い順、同数の場合は合計時間が多い順にシートを作成
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    for name, _ in sorted_users:
                        group = self.merged_df[self.merged_df['氏名'] == name]
                        group.to_excel(writer, sheet_name=name[:31], index=False)
                
                workbook = load_workbook(save_path)
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    
                    # セルの高さを27ピクセルに設定し、上下中央揃えを適用
                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                        worksheet.row_dimensions[row[0].row].height = 16
                        for cell in row:
                            cell.alignment = Alignment(vertical='center')
                    
                    # テーブルの範囲を定義
                    table_range = f"A1:{chr(64 + len(columns_to_output))}{worksheet.max_row}"
                    
                    # テーブルを作成
                    tab = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", 
                              ref=table_range)
                    
                    # テーブルスタイルを設定
                    tab.tableStyleInfo = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    
                    # テーブルをワークシートに追加
                    worksheet.add_table(tab)
                    
                    # フォルダ名ごとにマークXの有無とフォルダ内の修了証の有無を確認するための辞書
                    folder_has_x_mark = {}
                    folder_has_completion = {}
                    
                    # まず全行を走査してフォルダ名ごとにマークXの有無と修了証の有無を確認
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        folder_name = row[3].value  # 4列目が「フォルダ名」
                        mark = row[10].value  # 11列目が「マーク」
                        content_name = row[4].value  # 5列目が「コンテンツ名」
                        
                        if folder_name:
                            if not folder_name in folder_has_x_mark:
                                folder_has_x_mark[folder_name] = False
                                folder_has_completion[folder_name] = False
                            if mark == 'X':
                                folder_has_x_mark[folder_name] = True
                            if content_name and '修了証' in str(content_name):
                                folder_has_completion[folder_name] = True
                    
                    # 合計秒数をカウントする変数
                    total_seconds = 0 
                    
                    # 再度全行を走査して背景色を設定
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        content_name = row[4].value  # 5列目が「コンテンツ名」
                        folder_name = row[3].value  # 4列目が「フォルダ名」
                        confirm_time_cell = row[9]  # 10列目が「確認時間」

                        # 背景色の設定
                        should_be_orange = folder_name and (folder_has_x_mark[folder_name] or not folder_has_completion[folder_name])
                        should_be_yellow = content_name and '修了証' in str(content_name)

                        # 確認時間の集計（オレンジ背景でない場合のみ）
                        if not should_be_orange and confirm_time_cell.value:
                            total_seconds += self.time_to_seconds(confirm_time_cell.value)

                        # 背景色の適用
                        if should_be_yellow:
                            for cell in row:
                                cell.fill = yellow_fill
                        elif should_be_orange:
                            for cell in row:
                                cell.fill = orange_fill

                    # 確認時間の合計を最後のセルに表示（フォーマットを修正）
                    last_row = worksheet.max_row + 2  # テーブルの後に1行空けて表示
                    h = int(total_seconds // 3600)
                    m = int((total_seconds % 3600) // 60)
                    s = int(total_seconds % 60)
                    worksheet[f'J{last_row}'] = f"{h:02}:{m:02}:{s:02}"  # 10列目（J列）の最終行に合計時間を表示
                    
                    # 合計時間のセルに背景色を設定
                    total_hours = total_seconds / 3600  # 秒を時間に変換
                    if total_hours < 10:
                        worksheet[f'J{last_row}'].fill = red_fill
                    else:
                        worksheet[f'J{last_row}'].fill = green_fill

                workbook.save(save_path)
                print(json.dumps({
                    "status": "success",
                    "data": {
                        "message": f"処理が完了しました。ファイルは {save_path} に保存されました。",
                        "excel_path": save_path
                    },
                    "totalStandardTime": self.calculate_total_standard_time()
                }))
            else:
                print(json.dumps({
                    "status": "cancel",
                    "data": {
                        "message": "保存がキャンセルされました。計測結果は保存されていません。"
                    }
                }))
        except Exception as e:
            print(json.dumps({
                "status": "error",
                "data": {
                    "message": str(e)
                }
            }))

    def calculate_total_standard_time(self):
        total_seconds = 0
        for time_str in self.merged_df['標準学習時間']:
            total_seconds += self.time_to_seconds(time_str)
        return total_seconds

    def time_to_seconds(self, time_str):
        if pd.isna(time_str) or time_str == '':
            return 0
        h, m, s = map(int, time_str.split(':'))
        return h * 3600 + m * 60 + s

    def seconds_to_time(self, seconds):
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        s = int(seconds % 60)
        return f"{h:02}:{m:02}:{s:02}"

    def return_error_json(self, error_message):
        result = {
            'error': True,
            'message': error_message
        }
        print(json.dumps(result))
        sys.exit(1)

def main():
    processor = FileProcessor()
    processor.read_files()
    processor.process_data()
    processor.save_to_excel()

if __name__ == "__main__":
    main()
